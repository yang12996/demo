"""
封装的需求：
    1、数据读取：封装一个可以读取任意excel文件的方法，可以指定读取的表单
    2、数据写入：
          文件名：
           表单：
           行：
           列：
        写入值：
"""
import openpyxl

class HandleExcel:

   def __init__(self,filename,sheetname):
       """
       :param filename:excel文件名（路径）
       :param sheetname: 表单名
       """
       self.filename=filename
       self.sheetname=sheetname
       
   def read_data(self):
        """读取excel数据"""
        workbook=openpyxl.load_workbook(self.filename)
        sh=workbook[self.sheetname]
        res = list(sh.rows)
        #获取第一行的表头
        title =[i.value for i in res[0]]
        cases= []
        #遍历第一行之外的其他行
        for item in res[1:]:
             data=[i.value for i in item]
             dic = dict(zip(title,data))
             cases.append(dic)
        #返回读取出来的数据
        return  cases
   def write_data(self,row,column,value):
       """
       数据写入方法
       :param row:写入的行
       :param column:写入的列
       :param value:写入的值
       :return:
       """
       #加载工作簿对象
       workbook=openpyxl.load_workbook(self.filename)
       sh=workbook[self.sheetname]
       #写入数据
       sh.cell(row=row,column=column,value=value)
       workbook.save(self.filename)


if __name__ == '__main__':
    excel =HandleExcel(r"D:\Python\Unitest\demo2\datas\aa.xlsx", "login")
    res=excel.read_data()
    print(res)




